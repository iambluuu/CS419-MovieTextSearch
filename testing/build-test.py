
import os
import google.generativeai as genai
import pandas as pd
from openpyxl import Workbook, load_workbook
import time
import spacy

def extract_keywords(text):
    nlp = spacy.load("en_core_web_sm")
    doc = nlp(text)

    # Extract keywords (nouns, proper nouns, adjectives)
    keywords = ' '.join([token.text for token in doc if token.pos_ in ["NOUN", "PROPN", "ADJ"]])
    return keywords

genai.configure(api_key="AIzaSyBlE_sCA6L6hBCZPiIUWiaTyDLEx7Q5Mro")

import re

def sanitize_text(text):
    # Replace or remove potentially problematic content
    sanitized = re.sub(r"[^\w\s.,!?'-]", "", text)  # Remove special characters
    sanitized = re.sub(r"\b(prohibited|sensitive|explicit)\b", "[REDACTED]", sanitized, flags=re.IGNORECASE)
    return sanitized

# Create the model
generation_config = {
  "temperature": 1,
  "top_p": 0.95,
  "top_k": 40,
  "max_output_tokens": 8192,
  "response_mime_type": "text/plain",
}

SAFETY_SETTINGS = {
    genai.types.HarmCategory.HARM_CATEGORY_HATE_SPEECH: genai.types.HarmBlockThreshold.BLOCK_NONE,
    genai.types.HarmCategory.HARM_CATEGORY_DANGEROUS_CONTENT: genai.types.HarmBlockThreshold.BLOCK_NONE,
    genai.types.HarmCategory.HARM_CATEGORY_HARASSMENT: genai.types.HarmBlockThreshold.BLOCK_NONE,
    genai.types.HarmCategory.HARM_CATEGORY_SEXUALLY_EXPLICIT: genai.types.HarmBlockThreshold.BLOCK_NONE,
}

model = genai.GenerativeModel(
    model_name="gemini-1.5-flash",
    generation_config=generation_config,
    safety_settings=SAFETY_SETTINGS
)


# Function to summarize plot using Hugging Face Transformers
def summarize_plot(plot, max_length=50, max_retries=5):
    prompt = "Provide one plot point in the following synopsis in one sentence. Use only the words in the synopsis. Avoid offensive phrasing."
    sanitized_plot = sanitize_text(plot)

    try:
        response = model.generate_content([prompt, sanitized_plot])
        return response.text  # Adjust based on API response
    except Exception as e:
        if "PROHIBITED_CONTENT" in str(e):
            print(f"Blocked prompt detected. Skipping plot: {e}")
            return None
        
        if "429 Resource has been exhausted" in str(e):
            print("API rate limit reached. Please wait before retrying.")
            time.sleep(20)

        if max_retries > 0:
            print(f"Retrying to summarize plot. {max_retries} retries left.")
            time.sleep(1)
            return summarize_plot(plot, max_length, max_retries - 1)
        print(f"Failed to summarize plot: {e}")

        return None

def generate_keyword_queries(input_file, output_file, max_entries=1000):
    # Read the Excel file
    data = pd.read_excel(input_file)

    data = data.head(max_entries)
    print(data.columns)

    workbook = load_workbook(output_file) if os.path.exists(output_file) else Workbook()
    sheet = workbook.active

    for index, row in data.iterrows():
        title = row['title']
        print(f"{index}: Processing {title}...")
        summary = row['query']

        if not summary or not isinstance(summary, str) or len(summary) < 10:
            continue

        if not summary:
            continue
        keywords = extract_keywords(summary)
        sheet.cell(row=index+2, column=3, value=keywords)
    
    workbook.save(output_file)
    print(f"Generated queries have been saved to {output_file}")

def generate_queries_from_long_essays(input_file, output_file, max_entries=1000):
    # Read the Excel file
    data = pd.read_excel(input_file)

    # Ensure required columns are present
    if 'title' not in data.columns or 'plot_synopsis' not in data.columns:
        raise ValueError("Input file must contain 'title' and 'plot_synopsis' columns.")

    # Limit the data to the first `max_entries` rows
    data = data.head(max_entries)

    # Process each row to generate queries
    workbook = load_workbook(output_file) if os.path.exists(output_file) else Workbook()
    sheet = workbook.active
    current_row = sheet.max_row

    for index, row in data.iterrows():
        if index + 2 <= current_row:
            continue

        title = row['title']
        print(f"{index}: Processing {title}...")
        plot_synopsis = row['plot_synopsis']
        if isinstance(plot_synopsis, str):
            # Summarize the plot
            summary = summarize_plot(plot_synopsis)
            time.sleep(1)
            if summary:
                sheet.append([summary, title])
            else:
                sheet.append(["", title])
            workbook.save(output_file)

    print(f"Generated queries have been saved to {output_file}")

# Example usage
input_xlsx = "test_data.xlsx"   # Input file with 'title' and 'plot_synopsis' columns
output_xlsx = "test_data.xlsx"  # Output file to save generated queries
# generate_queries_from_long_essays(input_xlsx, output_xlsx, max_entries=1000)
generate_keyword_queries(input_xlsx, output_xlsx, max_entries=1000)

